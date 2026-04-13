"""
cpu_scheduler.py
Loads long/short/mixed workload CSVs, runs SJF, FCFS, HRRN on 4 cores,
and exports all results + summary to scheduler_results.xlsx

Algorithm behaviour:
  SJF  – Non-preemptive Shortest Job First: picks the ready process with the
          shortest known burst and runs it to completion without interruption.
  FCFS – Non-preemptive First Come First Served: picks the ready process with
          the earliest arrival time and runs it to completion without interruption.
  HRRN – Non-preemptive Highest Response Ratio Next: picks max (wait+burst)/burst.
"""

import csv
import os
import subprocess
import json
from dataclasses import dataclass
from typing import List, Dict, Optional
from enum import Enum

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Constants ──────────────────────────────────────────────────────────────────

NUM_CORES = 4

WORKLOAD_FILES = {
    "Long":  "long_processes.txt",
    "Short": "short_processes.txt",
    "Mixed": "mixed_processes.txt",
}
ALGORITHMS  = ["SJF", "FCFS", "HRRN"]
OUTPUT_FILE = "scheduler_results.xlsx"

# ── Data structures ────────────────────────────────────────────────────────────

@dataclass
class Process:
    name: str
    arrival: int
    burst: int
    start_time: int = -1
    finish_time: int = -1
    core: int = -1

    @property
    def turnaround(self): return self.finish_time - self.arrival
    @property
    def waiting(self):    return self.turnaround - self.burst

class Algorithm(Enum):
    SJF  = "SJF"
    FCFS = "FCFS"
    HRRN = "HRRN"

# ── CSV loader ─────────────────────────────────────────────────────────────────

def load_csv(filepath: str) -> List[Process]:
    procs = []
    with open(filepath, newline="") as f:
        reader = csv.DictReader(r for r in f if not r.startswith("#"))
        for row in reader:
            procs.append(Process(
                name    = row["name"].strip(),
                arrival = int(row["arrival"].strip()),
                burst   = int(row["burst"].strip()),
            ))
    return procs

# ── Non-preemptive scheduler (SJF / FCFS / HRRN) ──────────────────────────────

def simulate_nonpreemptive(procs: List[Process], algo: Algorithm) -> List[Process]:
    """
    Non-preemptive multi-core scheduler used by SJF, FCFS, and HRRN.
    Once a process is assigned to a core it runs to completion without interruption.
    Selection policy differs per algorithm:
      SJF  – min burst (ties broken by arrival, then name)
      FCFS – min arrival (ties broken by name)
      HRRN – max (wait + burst) / burst
    """
    remaining = [Process(p.name, p.arrival, p.burst) for p in procs]
    remaining.sort(key=lambda p: (p.arrival, p.name))
    completed  = []
    core_free  = [0] * NUM_CORES
    time       = 0

    while remaining:
        # Advance clock if nothing has arrived yet
        available = [p for p in remaining if p.arrival <= time]
        if not available:
            time = min(p.arrival for p in remaining)
            available = [p for p in remaining if p.arrival <= time]

        # Advance clock if no cores are free
        free_cores = [i for i, t in enumerate(core_free) if t <= time]
        if not free_cores:
            time = min(t for t in core_free if t > time)
            available = [p for p in remaining if p.arrival <= time]
            free_cores = [i for i, t in enumerate(core_free) if t <= time]

        for core_idx in free_cores:
            available = [p for p in remaining if p.arrival <= time]
            if not available:
                break

            if algo == Algorithm.SJF:
                chosen = min(available, key=lambda p: (p.burst, p.arrival, p.name))
            elif algo == Algorithm.FCFS:
                chosen = min(available, key=lambda p: (p.arrival, p.name))
            else:  # HRRN
                chosen = max(available, key=lambda p: (
                    ((time - p.arrival) + p.burst) / p.burst, -p.arrival))

            chosen.start_time   = time
            chosen.finish_time  = time + chosen.burst
            chosen.core         = core_idx
            core_free[core_idx] = chosen.finish_time
            remaining.remove(chosen)
            completed.append(chosen)

        if remaining:
            next_arrival   = min((p.arrival for p in remaining), default=float('inf'))
            busy_cores     = [t for t in core_free if t > time]
            next_core_free = min(busy_cores) if busy_cores else float('inf')
            time = min(next_arrival, next_core_free)

    return completed

# ── Unified dispatcher ─────────────────────────────────────────────────────────

def simulate(procs: List[Process], algo: Algorithm) -> List[Process]:
    return simulate_nonpreemptive(procs, algo)

# ── Metrics ────────────────────────────────────────────────────────────────────

def calc_metrics(procs: List[Process]) -> Dict:
    n     = len(procs)
    order = sorted(procs, key=lambda p: p.start_time)

    switches = 0
    for i in range(1, len(order)):
        switches += 1
        if order[i].start_time > order[i - 1].finish_time:
            switches += 1

    total_busy    = sum(p.burst for p in procs)
    first_arrival = min(p.arrival for p in procs)
    makespan      = max(p.finish_time for p in procs) - first_arrival
    cpu_util      = (total_busy / (makespan * NUM_CORES) * 100) if makespan > 0 else 0.0
    throughput    = n / makespan if makespan > 0 else 0.0

    core_procs    = {c: [p for p in procs if p.core == c] for c in range(NUM_CORES)}
    core_counts   = [len(core_procs[c]) for c in range(NUM_CORES)]
    core_busy     = [sum(p.burst for p in core_procs[c]) for c in range(NUM_CORES)]
    core_makespan = [
        (max(p.finish_time for p in core_procs[c]) - first_arrival)
        if core_procs[c] else 0
        for c in range(NUM_CORES)
    ]
    core_idle     = [core_makespan[c] - core_busy[c] for c in range(NUM_CORES)]

    return {
        "avg_turnaround":         sum(p.turnaround for p in procs) / n,
        "avg_waiting":            sum(p.waiting    for p in procs) / n,
        "avg_context_switches":   switches / n,
        "total_context_switches": switches,
        "cpu_utilization_pct":    cpu_util,
        "throughput":             throughput,
        "makespan":               makespan,
        "total_cpu_busy":         total_busy,
        "n":                      n,
        "avg_processes_per_core": n / NUM_CORES,
        "core_counts":            core_counts,
        "core_idle":              core_idle,
        "core_makespan":          core_makespan,
    }

# ── Shared style helpers ───────────────────────────────────────────────────────

DARK_NAVY   = "0D1B2A"
DARK_BLUE   = "1E3A5F"
MID_BLUE    = "2A5298"
TEAL        = "1A6B6B"
ROW_EVEN    = "0F2035"
ROW_ODD     = "152840"
AVG_GREEN   = "1A4A2E"
CORE_PURPLE = "2D1B5A"
CORE_VAL    = "3D2A7A"

_thin   = Side(style="thin", color="2A4A6A")
BORDER  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
CENTER  = Alignment(horizontal="center", vertical="center")
LEFT_A  = Alignment(horizontal="left",   vertical="center")

def _fill(c):           return PatternFill("solid", start_color=c)
def _row_fill(i):       return _fill(ROW_EVEN if i % 2 == 0 else ROW_ODD)
def _wf(sz=10, bd=True):return Font(color="FFFFFF", bold=bd, name="Arial", size=sz)
def _df(sz=10):         return Font(color="DDEEFF", bold=False, name="Arial", size=sz)
def _gf():              return Font(color="90EE90", bold=True,  name="Arial", size=10)
def _yf():              return Font(color="FFD700", bold=True,  name="Arial", size=10)
def _pf():              return Font(color="CF9FFF", bold=True,  name="Arial", size=10)

def _s(cell, font=None, fill=None, align=CENTER):
    if font:  cell.font      = font
    if fill:  cell.fill      = fill
    cell.alignment = align
    cell.border    = BORDER

def _hdr(ws, row, col, val, color=DARK_BLUE):
    c = ws.cell(row=row, column=col, value=val)
    _s(c, font=_wf(), fill=_fill(color))

def _title(ws, row, val, ncols, color=DARK_NAVY, sz=12):
    c = ws.cell(row=row, column=1, value=val)
    c.font      = Font(color="FFFFFF", bold=True, name="Arial", size=sz)
    c.fill      = _fill(color)
    c.alignment = LEFT_A
    c.border    = BORDER
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row,   end_column=ncols)
    ws.row_dimensions[row].height = 22

def _dat(ws, row, col, val, ri=0, fmt=None, font=None):
    c = ws.cell(row=row, column=col, value=val)
    _s(c, font=font or _df(), fill=_row_fill(ri))
    if fmt: c.number_format = fmt
    return c

# ── Detail sheet (one per workload/algorithm combo) ────────────────────────────

DCOLS   = ["Process","Arrival","Burst","Core","Start","Finish",
           "Turnaround","Waiting"]
DWIDTHS = [10, 10, 8, 8, 8, 10, 14, 10]

def build_detail_sheet(wb, name, procs, metrics, workload, algo):
    ws  = wb.create_sheet(name)
    n   = len(procs)
    NC  = len(DCOLS)

    _title(ws, 1,
           f"Workload: {workload}  │  Algorithm: {algo}  │  {n} Processes  │  {NUM_CORES} Cores",
           NC, sz=12)

    band = [
        ("Avg Turnaround",        f"{metrics['avg_turnaround']:.2f}"),
        ("Avg Waiting",           f"{metrics['avg_waiting']:.2f}"),
        ("Avg Ctx Switches/proc", f"{metrics['avg_context_switches']:.2f}"),
        ("CPU Utilization",       f"{metrics['cpu_utilization_pct']:.1f}%"),
        ("Throughput (p/tick)",   f"{metrics['throughput']:.5f}"),
        ("Makespan",              f"{metrics['makespan']}"),
        ("Total Ctx Switches",    f"{metrics['total_context_switches']}"),
        ("Avg Proc/Core",         f"{metrics['avg_processes_per_core']:.2f}"),
    ]
    for idx, (lbl, val) in enumerate(band):
        col = (idx % 4) * 2 + 1
        row = 2 + idx // 4
        c = ws.cell(row=row, column=col, value=lbl)
        _s(c, font=_wf(sz=9), fill=_fill(MID_BLUE))
        c = ws.cell(row=row, column=col+1, value=val)
        _s(c, font=_yf(), fill=_fill(TEAL))

    core_row = 4
    core_labels = [""] + [f"Core {c}" for c in range(NUM_CORES)]
    core_metric_rows = [
        ("Proc Count", metrics["core_counts"]),
        ("Idle Time",  metrics["core_idle"]),
        ("Makespan",   metrics["core_makespan"]),
    ]
    for ci, lbl in enumerate(core_labels):
        c = ws.cell(row=core_row, column=ci + 1, value=lbl)
        _s(c, font=_wf(sz=9), fill=_fill(CORE_PURPLE))

    for ri, (metric_lbl, values) in enumerate(core_metric_rows):
        r = core_row + 1 + ri
        c = ws.cell(row=r, column=1, value=metric_lbl)
        _s(c, font=_wf(sz=9), fill=_fill(CORE_PURPLE))
        for ci, val in enumerate(values):
            c = ws.cell(row=r, column=ci + 2, value=val)
            _s(c, font=_pf(), fill=_fill(CORE_VAL))

    HDR = core_row + 1 + len(core_metric_rows) + 1
    for ci, lbl in enumerate(DCOLS, 1):
        _hdr(ws, HDR, ci, lbl)
    ws.row_dimensions[HDR].height = 18

    order      = sorted(procs, key=lambda p: p.start_time)
    first_data = HDR + 1
    for ri, p in enumerate(order):
        r = first_data + ri
        for ci, val in enumerate(
            [f"P{p.name}", p.arrival, p.burst, f"Core {p.core}",
             p.start_time, p.finish_time, p.turnaround, p.waiting], 1):
            _dat(ws, r, ci, val, ri=ri)

    last_data = first_data + n - 1

    AR = last_data + 1
    for ci in range(1, NC + 1):
        col_l = get_column_letter(ci)
        if ci == 1:
            c = ws.cell(row=AR, column=ci, value="AVERAGE")
        elif ci in (5, 6):
            c = ws.cell(row=AR, column=ci, value="")
        elif ci in (7, 8):
            c = ws.cell(row=AR, column=ci,
                        value=f"=AVERAGE({col_l}{first_data}:{col_l}{last_data})")
            c.number_format = "0.00"
        else:
            c = ws.cell(row=AR, column=ci, value="")
        _s(c, font=_gf(), fill=_fill(AVG_GREEN))

    for i, w in enumerate(DWIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{first_data}"

# ── Summary sheet ──────────────────────────────────────────────────────────────

SCOLS = [
    "Workload","Algorithm","Processes","Makespan",
    "Avg Turnaround","Avg Waiting",
    "Avg Ctx Sw/Proc","Total Ctx Switches",
    "CPU Util %","Throughput (p/tick)","Avg Proc/Core",
    "C0 Procs","C1 Procs","C2 Procs","C3 Procs",
    "C0 Idle", "C1 Idle", "C2 Idle", "C3 Idle",
    "C0 Makespan","C1 Makespan","C2 Makespan","C3 Makespan",
]
SWIDTHS = [12, 12, 12, 12, 17, 15, 18, 20, 14, 22, 15,
           10, 10, 10, 10,
           10, 10, 10, 10,
           13, 13, 13, 13]
SFMTS   = [None,None,None,None,"0.00","0.00","0.00",None,"0.0","0.0000","0.00",
           None,None,None,None,
           None,None,None,None,
           None,None,None,None]

def build_summary_sheet(wb, all_results):
    ws   = wb.create_sheet("Summary", 0)
    NC   = len(SCOLS)

    _title(ws, 1, f"CPU Scheduling — Full Summary  ({NUM_CORES} Cores)", NC, sz=13)

    for ci, lbl in enumerate(SCOLS, 1):
        color = CORE_PURPLE if ci > 11 else DARK_BLUE
        _hdr(ws, 2, ci, lbl, color=color)
    ws.row_dimensions[2].height = 18

    for ri, (wl, algo, _, m) in enumerate(all_results):
        r = 3 + ri
        row_data = [
            wl, algo, m["n"], m["makespan"],
            m["avg_turnaround"], m["avg_waiting"],
            m["avg_context_switches"], m["total_context_switches"],
            m["cpu_utilization_pct"], m["throughput"],
            m["avg_processes_per_core"],
            *m["core_counts"],
            *m["core_idle"],
            *m["core_makespan"],
        ]
        for ci, (val, fmt) in enumerate(zip(row_data, SFMTS), 1):
            font = _pf() if ci > 11 else None
            _dat(ws, r, ci, val, ri=ri, fmt=fmt, font=font)

    last_data = 2 + len(all_results)

    sep = last_data + 2
    _title(ws, sep, "Per-Workload Averages  (across all 3 algorithms)", NC,
           color=MID_BLUE, sz=11)
    for ci, lbl in enumerate(SCOLS, 1):
        color = CORE_PURPLE if ci > 11 else TEAL
        _hdr(ws, sep + 1, ci, lbl, color=color)

    workloads = list(dict.fromkeys(r[0] for r in all_results))
    for wi, wl in enumerate(workloads):
        excel_rows = [3 + i for i, r in enumerate(all_results) if r[0] == wl]
        r = sep + 2 + wi

        c = ws.cell(row=r, column=1, value=wl)
        _s(c, font=_wf(), fill=_fill(DARK_BLUE))
        c = ws.cell(row=r, column=2, value="All algos")
        _s(c, font=_df(), fill=_row_fill(wi))

        for ci in range(3, NC + 1):
            col_l = get_column_letter(ci)
            refs  = ",".join(f"{col_l}{er}" for er in excel_rows)
            c = ws.cell(row=r, column=ci, value=f"=AVERAGE({refs})")
            font = _pf() if ci > 11 else _gf()
            _s(c, font=font, fill=_fill(AVG_GREEN))
            fmt = SFMTS[ci - 1]
            if fmt: c.number_format = fmt

    for i, w in enumerate(SWIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"

# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    all_results = []

    for wl_name, filename in WORKLOAD_FILES.items():
        if not os.path.exists(filename):
            print(f"  [SKIP] {filename} not found")
            continue
        procs = load_csv(filename)
        print(f"\nWorkload: {wl_name}  ({len(procs)} processes)  [{NUM_CORES} cores]")

        for algo_name in ALGORITHMS:
            results = simulate(procs, Algorithm[algo_name])
            metrics = calc_metrics(results)
            sheet_name = f"{wl_name[:5]}-{algo_name}"
            build_detail_sheet(wb, sheet_name, results, metrics, wl_name, algo_name)
            all_results.append((wl_name, algo_name, results, metrics))

            print(f"  {algo_name}  TA={metrics['avg_turnaround']:8.2f}  "
                  f"WT={metrics['avg_waiting']:8.2f}  "
                  f"CS/p={metrics['avg_context_switches']:5.2f}  "
                  f"CPU={metrics['cpu_utilization_pct']:5.1f}%  "
                  f"P/Core={metrics['avg_processes_per_core']:.2f}  "
                  f"Counts={metrics['core_counts']}  "
                  f"Idle={metrics['core_idle']}")

    build_summary_sheet(wb, all_results)
    wb.save(OUTPUT_FILE)
    print(f"\n  Saved → {OUTPUT_FILE}")

    # Recalculate formulas via LibreOffice
    res = subprocess.run(
        ["python", "scripts/recalc.py", OUTPUT_FILE],
        capture_output=True, text=True
    )
    try:
        info = json.loads(res.stdout)
        status = info.get("status", "?")
        print(f"  Recalc: {status}  |  formulas={info.get('total_formulas',0)}  "
              f"errors={info.get('total_errors',0)}")
        if info.get("error_summary"):
            print(f"  Errors: {info['error_summary']}")
    except Exception:
        print(f"  Recalc: {res.stdout.strip() or res.stderr.strip()}")

if __name__ == "__main__":
    main()